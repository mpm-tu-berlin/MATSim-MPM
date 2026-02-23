"""
Fragestellung:
  Wie viele Fahrzeuge der Gruppen 5, 10 und 12 mit Eintrag in PEVArchitecture
  haben Eintraege in der Achsdatei (hdv_2023_axle.csv)?

Strategie (lokal effizient):
  1. Axle-CSV einmal vollstaendig laden -> set fuer O(1)-Lookup.
  2. Vehicle-CSV einmal chunked durchlaufen, alle drei Gruppen gleichzeitig filtern.
  3. Join im RAM, kein Netz, kein DB.
  4. Excel: 1 Zusammenfassungs-Sheet + je 2 Sheets pro Gruppe.
"""

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# === Konfiguration ===
GRUPPEN      = ["5", "10", "12"]
PROJECT_ROOT = Path(__file__).resolve().parent.parent
EEA_DIR      = PROJECT_ROOT / "data" / "eea_t_co2-emission-hdv_p_2023-2024_v01_r00"
VEHICLE_CSV  = EEA_DIR / "hdv_2023_vehicle.csv"
AXLE_CSV     = EEA_DIR / "hdv_2023_axle.csv"
OUTPUT_FILE  = PROJECT_ROOT / "results" / "gruppen_pev_axle.xlsx"

VEHICLE_COLS = [
    "Vehicle_id", "Make", "Model", "VehicleGroup",
    "PEVArchitecture", "HEVArchitecture",
    "ZeroEmissionVehicle", "AxleConfiguration",
    "CorrectedActualMass", "TechnicalPermissibleMaximumLadenMass",
    "SumNetPower",
]

CSV_OPTS = dict(sep=";", dtype=str, encoding="utf-8",
                on_bad_lines="skip", quotechar='"')

# Farben pro Gruppe fuer die Sheet-Header
GRUPPE_FARBEN = {"5": "1F4E79", "10": "375623", "12": "7B2C2C"}


# =============================================================================
# Hilfsfunktionen
# =============================================================================

def style_header(ws, row: int = 1, color: str = "1F4E79"):
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    for cell in ws[row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def autofit(ws, max_w: int = 55):
    for col in ws.columns:
        width = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 2, max_w)


# =============================================================================
# 1. Axle-CSV einlesen
# =============================================================================
print("Lade Achsdaten …")
axle_df = pd.read_csv(AXLE_CSV, **CSV_OPTS)
axle_df["Vehicle_id"] = axle_df["Vehicle_id"].str.strip('"')
axle_ids: set = set(axle_df["Vehicle_id"].dropna())
print(f"  {len(axle_ids):,} eindeutige Fahrzeuge in Achsdatei.")

# =============================================================================
# 2. Vehicle-CSV einmal durchlaufen, alle Gruppen gleichzeitig filtern
# =============================================================================
print(f"Filtere Vehicle-CSV (Gruppen {', '.join(GRUPPEN)}, PEVArchitecture gefuellt) …")

frames = []
for chunk in pd.read_csv(VEHICLE_CSV, usecols=VEHICLE_COLS,
                         chunksize=100_000, **CSV_OPTS):
    for col in ("Vehicle_id", "VehicleGroup", "PEVArchitecture"):
        chunk[col] = chunk[col].str.strip('"')

    mask = (
        chunk["VehicleGroup"].isin(GRUPPEN) &
        chunk["PEVArchitecture"].notna() &
        (chunk["PEVArchitecture"] != "")
    )
    frames.append(chunk[mask])

all_veh = pd.concat(frames, ignore_index=True)

# =============================================================================
# 3. Axle-Join
# =============================================================================
all_veh["hat_Achsdaten"] = all_veh["Vehicle_id"].isin(axle_ids)

# Ergebnisse je Gruppe berechnen
stats = {}
for g in GRUPPEN:
    gdf = all_veh[all_veh["VehicleGroup"] == g]
    n = len(gdf)
    m = int(gdf["hat_Achsdaten"].sum())
    stats[g] = {"df": gdf, "gesamt": n, "mit": m, "ohne": n - m}
    print(f"  Gruppe {g:>2}: {n:>4} Fahrzeuge | {m:>4} mit Achsdaten | {n-m:>4} ohne")

# =============================================================================
# 4. Excel schreiben
# =============================================================================
OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
wb = Workbook()
del wb[wb.sheetnames[0]]

# --- Sheet 1: Zusammenfassung aller Gruppen ---
ws_sum = wb.create_sheet("Zusammenfassung")
ws_sum.append(["Gruppe", "Fahrzeuge mit PEVArchitecture",
               "davon mit Achsdaten", "davon ohne Achsdaten", "Anteil mit Achsdaten"])
style_header(ws_sum)

for g in GRUPPEN:
    s = stats[g]
    anteil = f"{s['mit']/s['gesamt']*100:.1f} %" if s["gesamt"] else "–"
    ws_sum.append([f"Gruppe {g}", s["gesamt"], s["mit"], s["ohne"], anteil])

ws_sum.append([])
ws_sum.append(["Quelldateien"])
ws_sum.append(["Vehicle-CSV", str(VEHICLE_CSV.relative_to(PROJECT_ROOT))])
ws_sum.append(["Axle-CSV",    str(AXLE_CSV.relative_to(PROJECT_ROOT))])
autofit(ws_sum)

# --- Je Gruppe: 2 Sheets ---
display_cols = [c for c in VEHICLE_COLS if c in all_veh.columns] + ["hat_Achsdaten"]

for g in GRUPPEN:
    s = stats[g]
    farbe = GRUPPE_FARBEN[g]
    gdf = s["df"]

    # Sheet A: Alle Fahrzeuge der Gruppe
    ws_all = wb.create_sheet(f"Gruppe{g} alle PEV")
    ws_all.append(display_cols)
    style_header(ws_all, color=farbe)
    for _, row in gdf[display_cols].sort_values("hat_Achsdaten", ascending=False).iterrows():
        ws_all.append([str(v) if pd.notna(v) else "" for v in row])
    autofit(ws_all)

    # Sheet B: Nur Matches mit Achsdetails
    ws_match = wb.create_sheet(f"Gruppe{g} Matches")
    matches = gdf[gdf["hat_Achsdaten"]]
    merged = matches.merge(axle_df, on="Vehicle_id", how="left",
                           suffixes=("_veh", "_axle"))
    ws_match.append(list(merged.columns))
    style_header(ws_match, color=farbe)
    for _, row in merged.iterrows():
        ws_match.append([str(v) if pd.notna(v) else "" for v in row])
    autofit(ws_match)

wb.save(OUTPUT_FILE)
print(f"\nFertig! Excel gespeichert: {OUTPUT_FILE}")

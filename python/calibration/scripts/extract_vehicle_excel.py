"""
Extrahiert alle Daten fuer ein einzelnes Fahrzeug aus den EEA-HDV-2023-Quellen
und schreibt sie in eine Excel-Datei mit je einem Sheet pro Quelldatei.

Aufruf:
    python scripts/extract_vehicle_excel.py

Die Fahrzeug-ID und der Ausgabepfad sind unten konfigurierbar.
"""

import csv
import zipfile
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# === Konfiguration ===
VEHICLE_ID = "0x82733F217054FA16F4E3434BF4D4BE861B014FE6A2F341C65B0B58AB224E05A7"

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = PROJECT_ROOT / "data"
EEA_DIR = DATA_DIR / "eea_t_co2-emission-hdv_p_2023-2024_v01_r00"

OUTPUT_FILE = PROJECT_ROOT / "results" / f"fahrzeug_daten_{VEHICLE_ID[:10]}.xlsx"

# Quelldateien
VEHICLE_CSV      = EEA_DIR / "hdv_2023_vehicle.csv"
AXLE_CSV         = EEA_DIR / "hdv_2023_axle.csv"
MISSION_CSV      = DATA_DIR / "hdv_2023_missionprofile" / "hdv_2023_missionprofile.csv"
REFERENCE_CSV    = DATA_DIR / "reference_consumption.csv"

ZIP_SOURCES = {
    "Batterie":           EEA_DIR / "hdv_2023_battery.zip",
    "Elektromotor":       EEA_DIR / "hdv_2023_electricmachine.zip",
    "Kraftstofftyp":      EEA_DIR / "hdv_2023_enginefueltype.zip",
    "Motordrehmoment":    EEA_DIR / "hdv_2023_enginetorquelimit.zip",
    "IEPC":               EEA_DIR / "hdv_2023_iepc.zip",
    "Lenkpumpe":          EEA_DIR / "hdv_2023_steeringpumptechnology.zip",
}


# =============================================================================
# Hilfsfunktionen
# =============================================================================

def style_header_row(ws, row_idx: int):
    """Formatiert eine Kopfzeile: fett, grauer Hintergrund."""
    fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    for cell in ws[row_idx]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True)


def autofit_columns(ws, max_width: int = 60):
    """Setzt Spaltenbreiten automatisch (begrenzt auf max_width)."""
    for col_cells in ws.columns:
        max_len = max(
            (len(str(c.value)) if c.value is not None else 0) for c in col_cells
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(
            max_len + 2, max_width
        )


def csv_row_to_dict(csv_path: Path, vehicle_id: str) -> dict | None:
    """Liest eine grosse CSV (semikolon-getrennt) zeilenweise und gibt die erste
    Zeile zurueck, deren Vehicle_id/vehicle_id-Spalte mit vehicle_id uebereinstimmt.
    Gibt None zurueck, wenn die ID nicht gefunden wird.
    """
    with open(csv_path, encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f, delimiter=";")
        for row in reader:
            vid = row.get("Vehicle_id") or row.get("vehicle_id") or ""
            # Anführungszeichen entfernen (manche Felder sind in "" eingeschlossen)
            if vid.strip('"') == vehicle_id:
                return {k: v.strip('"') for k, v in row.items()}
    return None


def csv_rows_to_df(csv_path: Path, vehicle_id: str) -> pd.DataFrame:
    """Liest alle Zeilen einer grossen CSV fuer eine bestimmte Fahrzeug-ID
    per Chunks (speichereffizient) und gibt einen DataFrame zurueck."""
    frames = []
    for chunk in pd.read_csv(csv_path, sep=";", chunksize=50_000,
                             dtype=str, encoding="utf-8", on_bad_lines="skip"):
        # Ersten Spaltennamen normalisieren (kann 'vehicle_id' oder 'Vehicle_id' sein)
        id_col = [c for c in chunk.columns if c.lower() == "vehicle_id"]
        if not id_col:
            continue
        col = id_col[0]
        mask = chunk[col].str.strip('"') == vehicle_id
        hits = chunk[mask]
        if not hits.empty:
            frames.append(hits)
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()


def zip_csv_rows_to_df(zip_path: Path, vehicle_id: str) -> pd.DataFrame:
    """Liest CSV-Dateien aus einem ZIP-Archiv und filtert nach Fahrzeug-ID."""
    if not zip_path.exists():
        return pd.DataFrame()
    frames = []
    with zipfile.ZipFile(zip_path) as zf:
        for name in zf.namelist():
            if not name.lower().endswith(".csv"):
                continue
            with zf.open(name) as f:
                try:
                    df = pd.read_csv(f, sep=";", dtype=str, encoding="utf-8",
                                     on_bad_lines="skip")
                except Exception:
                    continue
                id_col = [c for c in df.columns if c.lower() == "vehicle_id"]
                if not id_col:
                    continue
                col = id_col[0]
                mask = df[col].str.strip('"') == vehicle_id
                hits = df[mask]
                if not hits.empty:
                    frames.append(hits)
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()


# =============================================================================
# Sheet-Schreiber
# =============================================================================

def write_uebersicht(wb, sheets_info: list[dict]):
    """Schreibt ein Uebersichts-Sheet: Tabellenname, Quelldatei, Zeilenanzahl."""
    ws = wb.create_sheet("Uebersicht", 0)
    headers = ["Sheet", "Quelldatei", "Beschreibung", "Zeilen gefunden"]
    ws.append(headers)
    style_header_row(ws, 1)
    for info in sheets_info:
        ws.append([info["sheet"], info["quelle"], info["beschreibung"], info["zeilen"]])
    ws.append([])
    ws.append(["Fahrzeug-ID:", VEHICLE_ID])
    ws["A" + str(ws.max_row)].font = Font(bold=True)
    autofit_columns(ws)


def write_vehicle_sheet(wb, row: dict | None) -> dict:
    """Schreibt Fahrzeug-Stammdaten (hdv_2023_vehicle.csv) als Spalte -> Wert."""
    ws = wb.create_sheet("Fahrzeugdaten")
    ws.append(["Spalte", "Wert", "Quelldatei"])
    style_header_row(ws, 1)
    zeilen = 0
    if row:
        for col, val in row.items():
            ws.append([col, val, str(VEHICLE_CSV.relative_to(PROJECT_ROOT))])
            zeilen += 1
    autofit_columns(ws)
    return {"sheet": "Fahrzeugdaten", "quelle": str(VEHICLE_CSV.name),
            "beschreibung": "Fahrzeug-Stammdaten (Hersteller, Motor, Getriebe, …)",
            "zeilen": zeilen}


def write_axle_sheet(wb, df: pd.DataFrame) -> dict:
    """Schreibt Achsdaten (hdv_2023_axle.csv)."""
    ws = wb.create_sheet("Achsen")
    if df.empty:
        ws.append(["Keine Daten fuer diese Fahrzeug-ID gefunden."])
        return {"sheet": "Achsen", "quelle": str(AXLE_CSV.name),
                "beschreibung": "Achsen / Reifenspezifikation", "zeilen": 0}
    headers = list(df.columns) + ["Quelldatei"]
    ws.append(headers)
    style_header_row(ws, 1)
    for _, r in df.iterrows():
        ws.append(list(r) + [str(AXLE_CSV.relative_to(PROJECT_ROOT))])
    autofit_columns(ws)
    return {"sheet": "Achsen", "quelle": str(AXLE_CSV.name),
            "beschreibung": "Achsen / Reifenspezifikation", "zeilen": len(df)}


def write_mission_sheet(wb, df: pd.DataFrame) -> dict:
    """Schreibt Mission-Profile-Daten (hdv_2023_missionprofile.csv)."""
    ws = wb.create_sheet("Missionsprofil")
    if df.empty:
        ws.append(["Keine Daten fuer diese Fahrzeug-ID gefunden."])
        return {"sheet": "Missionsprofil", "quelle": str(MISSION_CSV.name),
                "beschreibung": "Simulationsergebnisse je Mission/Modus", "zeilen": 0}
    headers = list(df.columns) + ["Quelldatei"]
    ws.append(headers)
    style_header_row(ws, 1)
    for _, r in df.iterrows():
        ws.append(list(r) + [str(MISSION_CSV.relative_to(PROJECT_ROOT))])
    autofit_columns(ws)
    return {"sheet": "Missionsprofil", "quelle": str(MISSION_CSV.name),
            "beschreibung": "Simulationsergebnisse je Mission/Modus", "zeilen": len(df)}


def write_reference_sheet(wb, vehicle_group: str | None) -> dict:
    """Schreibt die passenden Referenzverbrauchswerte aus reference_consumption.csv."""
    ws = wb.create_sheet("Referenzverbrauch")
    ws.append(["vehicle_id", "vehicle_group", "mission", "payload_kg",
               "ee_kwh_per_km", "route_km", "n_vehicles",
               "Hinweis", "Quelldatei"])
    style_header_row(ws, 1)

    zeilen = 0
    hinweis = (
        f"Medianwert fuer Fahrzeuggruppe {vehicle_group}"
        if vehicle_group else "Fahrzeuggruppe unbekannt"
    )
    quelle_rel = str(REFERENCE_CSV.relative_to(PROJECT_ROOT))

    with open(REFERENCE_CSV, encoding="utf-8", newline="") as f:
        reader = csv.DictReader(
            (row for row in f if not row.startswith("#"))
        )
        for row in reader:
            if vehicle_group and str(row.get("vehicle_group", "")) != vehicle_group:
                continue
            ws.append([
                row.get("vehicle_id"), row.get("vehicle_group"),
                row.get("mission"), row.get("payload_kg"),
                row.get("ee_kwh_per_km"), row.get("route_km"),
                row.get("n_vehicles"), hinweis, quelle_rel,
            ])
            zeilen += 1

    if zeilen == 0:
        # Kein Filter moeglich -> alle laden
        with open(REFERENCE_CSV, encoding="utf-8", newline="") as f:
            reader = csv.DictReader(
                (row for row in f if not row.startswith("#"))
            )
            for row in reader:
                ws.append([
                    row.get("vehicle_id"), row.get("vehicle_group"),
                    row.get("mission"), row.get("payload_kg"),
                    row.get("ee_kwh_per_km"), row.get("route_km"),
                    row.get("n_vehicles"), "Alle Gruppen (kein Match)", quelle_rel,
                ])
                zeilen += 1

    autofit_columns(ws)
    return {"sheet": "Referenzverbrauch", "quelle": str(REFERENCE_CSV.name),
            "beschreibung": "VECTO-Referenzverbrauch (Medianwert der Fahrzeuggruppe)",
            "zeilen": zeilen}


def write_zip_sheet(wb, label: str, zip_path: Path, df: pd.DataFrame) -> dict:
    """Schreibt Daten aus einer ZIP-Quelldatei."""
    ws = wb.create_sheet(label[:31])  # Sheet-Name max 31 Zeichen
    if df.empty:
        ws.append([f"Keine Daten fuer diese Fahrzeug-ID in {zip_path.name}."])
        return {"sheet": label, "quelle": str(zip_path.name),
                "beschreibung": label, "zeilen": 0}
    headers = list(df.columns) + ["Quelldatei"]
    ws.append(headers)
    style_header_row(ws, 1)
    for _, r in df.iterrows():
        ws.append(list(r) + [zip_path.name])
    autofit_columns(ws)
    return {"sheet": label, "quelle": str(zip_path.name),
            "beschreibung": label, "zeilen": len(df)}


# =============================================================================
# Hauptprogramm
# =============================================================================

def main():
    print(f"Extrahiere Daten fuer Fahrzeug-ID: {VEHICLE_ID}")
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)

    # --- 1. Fahrzeug-Stammdaten ---
    print("  [1/6] Lese hdv_2023_vehicle.csv …")
    vehicle_row = csv_row_to_dict(VEHICLE_CSV, VEHICLE_ID)
    if vehicle_row:
        print(f"        Gefunden: {vehicle_row.get('Make', '')} {vehicle_row.get('Model', '')}")
        vehicle_group = vehicle_row.get("VehicleGroup") or vehicle_row.get("VehicleGroupCO2") or None
    else:
        print("        WARNUNG: Fahrzeug-ID nicht in hdv_2023_vehicle.csv gefunden.")
        vehicle_group = None

    # --- 2. Achsdaten ---
    print("  [2/6] Lese hdv_2023_axle.csv …")
    axle_df = csv_rows_to_df(AXLE_CSV, VEHICLE_ID)
    print(f"        {len(axle_df)} Achszeile(n) gefunden.")

    # --- 3. Missionsprofil ---
    print("  [3/6] Lese hdv_2023_missionprofile.csv (kann etwas dauern) …")
    mission_df = csv_rows_to_df(MISSION_CSV, VEHICLE_ID)
    print(f"        {len(mission_df)} Missionszeile(n) gefunden.")

    # --- 4. ZIP-Quellen ---
    zip_dfs = {}
    for i, (label, zip_path) in enumerate(ZIP_SOURCES.items(), start=4):
        print(f"  [{i}/{3 + len(ZIP_SOURCES)}] Lese {zip_path.name} …")
        zip_dfs[label] = zip_csv_rows_to_df(zip_path, VEHICLE_ID)
        print(f"        {len(zip_dfs[label])} Zeile(n) gefunden.")

    # --- Excel erstellen ---
    print("  Schreibe Excel …")
    from openpyxl import Workbook
    wb = Workbook()
    # Standard-Sheet entfernen
    del wb[wb.sheetnames[0]]

    sheets_info = []
    sheets_info.append(write_vehicle_sheet(wb, vehicle_row))
    sheets_info.append(write_axle_sheet(wb, axle_df))
    sheets_info.append(write_mission_sheet(wb, mission_df))
    sheets_info.append(write_reference_sheet(wb, vehicle_group))
    for label, df in zip_dfs.items():
        sheets_info.append(write_zip_sheet(wb, label, ZIP_SOURCES[label], df))

    # Uebersicht vorne einfuegen
    write_uebersicht(wb, sheets_info)

    wb.save(OUTPUT_FILE)
    print(f"\nFertig! Excel gespeichert unter:\n  {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
